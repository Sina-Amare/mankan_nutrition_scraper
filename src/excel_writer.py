"""Excel writer with professional styling for nutritional data."""

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from src.logger_config import get_logger

logger = get_logger(__name__)


class ExcelWriter:
    """Writes nutritional data to styled Excel file."""
    
    # Column definitions
    COLUMNS = [
        ("food_name", "Food Name"),
        ("measurement_unit", "Measurement Unit"),
        ("measurement_value", "Measurement Value"),
        ("calories", "Calories"),
        ("carbs_g", "Carbs (g)"),
        ("protein_g", "Protein (g)"),
        ("fat_g", "Fat (g)"),
        ("fiber_g", "Fiber (g)"),
        ("food_id", "Food ID"),
    ]
    
    # Styling constants
    HEADER_FILL = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid"
    )
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center")
    
    def __init__(self, output_dir: Path = Path("output")):
        """Initialize Excel writer.
        
        Args:
            output_dir: Directory for output files
        """
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def write_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str = "mankan_nutritional_data.xlsx"
    ) -> Path:
        """Write data to styled Excel file.
        
        Args:
            data: List of data dictionaries
            filename: Output filename
        Returns:
            Path to created Excel file
        """
        if not data:
            logger.warning("No data to write to Excel")
            return None
        
        output_path = self.output_dir / filename
        
        logger.info(f"Writing {len(data)} rows to Excel: {output_path}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Nutritional Data"
        
        # Write headers
        header_row = 1
        for col_idx, (field, header) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.ALIGNMENT_CENTER
            cell.border = self.BORDER
        
        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, (field, _) in enumerate(self.COLUMNS, start=1):
                value = row_data.get(field)
                # Convert None to empty string for Excel
                if value is None:
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                
                # Set alignment based on column type
                if field in ["food_name", "measurement_unit", "source_url"]:
                    cell.alignment = self.ALIGNMENT_LEFT
                else:
                    cell.alignment = self.ALIGNMENT_CENTER
        
        # Auto-adjust column widths
        for col_idx, (field, header) in enumerate(self.COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            
            # Calculate max width
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                if row[0].value:
                    max_length = max(max_length, len(str(row[0].value)))
            
            # Set width with some padding
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Create summary sheet
        self._create_summary_sheet(wb, data)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel file saved: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, wb: Workbook, data: List[Dict[str, Any]]):
        """Create summary statistics sheet.
        
        Args:
            wb: Workbook object
            data: List of data dictionaries
        """
        ws = wb.create_sheet("Summary", 0)
        
        # Calculate statistics
        unique_foods = len(set(row.get("food_id") for row in data))
        total_rows = len(data)
        completion_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Get measurement unit distribution
        measurement_counts = {}
        for row in data:
            unit = row.get("measurement_unit", "Unknown")
            measurement_counts[unit] = measurement_counts.get(unit, 0) + 1
        
        # Write summary
        summary_data = [
            ["Mankan.me Nutritional Database - Summary", ""],
            ["", ""],
            ["Completion Date", completion_date],
            ["Total Food Items", unique_foods],
            ["Total Data Rows", total_rows],
            ["Average Measurements per Food", f"{total_rows / unique_foods:.2f}" if unique_foods > 0 else "0"],
            ["", ""],
            ["Measurement Unit Distribution", ""],
        ]
        
        # Sort measurement units by count
        sorted_units = sorted(measurement_counts.items(), key=lambda x: x[1], reverse=True)
        for unit, count in sorted_units[:10]:  # Top 10
            summary_data.append([unit, count])
        
        # Write to sheet
        for row_idx, row_data in enumerate(summary_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Title row
                    cell.font = Font(bold=True, size=14)
                elif row_idx <= 7:  # Stats rows
                    if col_idx == 1:
                        cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
    
    def write_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str = "mankan_nutritional_data.csv"
    ) -> Path:
        """Write data to CSV file as backup.
        
        Args:
            data: List of data dictionaries
            filename: Output filename
        Returns:
            Path to created CSV file
        """
        if not data:
            logger.warning("No data to write to CSV")
            return None
        
        output_path = self.output_dir / filename
        
        logger.info(f"Writing {len(data)} rows to CSV: {output_path}")
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Reorder columns to match Excel
        column_order = [field for field, _ in self.COLUMNS]
        df = df.reindex(columns=column_order)
        
        # Write CSV
        df.to_csv(output_path, index=False, encoding="utf-8-sig")
        
        logger.info(f"CSV file saved: {output_path}")
        
        return output_path

