"""Incremental writer for CSV and Excel files with batch appending."""

import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from src.logger_config import get_logger

logger = get_logger(__name__)


class IncrementalWriter:
    """Handles incremental/batch writing to CSV and Excel files."""
    
    # Column definitions - matching image format order
    COLUMNS = [
        ("food_name", "Food Name"),           # نام
        ("measurement_unit", "Measurement Unit"),  # واحد
        ("calories", "Calories"),             # کالری
        ("fat_g", "Fat (g)"),                 # چربی
        ("protein_g", "Protein (g)"),         # پروتئین
        ("carbs_g", "Carbs (g)"),             # کربوهیدرات
        ("fiber_g", "Fiber (g)"),             # فیبر
        ("sugar_g", "Sugar (g)"),             # قند
        ("salt_g", "Salt (g)"),               # نمک
        ("measurement_value", "Measurement Value"),  # Internal
        ("food_id", "Food ID"),               # Internal
    ]
    
    # Styling constants
    HEADER_FILL = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid"
    )
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center")
    
    def __init__(
        self,
        output_dir: Path = Path("output"),
        csv_filename: str = "mankan_nutritional_data.csv",
        excel_filename: str = "mankan_nutritional_data.xlsx",
        batch_size: int = 50
    ):
        """Initialize incremental writer.
        
        Args:
            output_dir: Directory for output files
            csv_filename: CSV output filename
            excel_filename: Excel output filename
            batch_size: Number of items to accumulate before writing
        """
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.csv_path = self.output_dir / csv_filename
        self.excel_path = self.output_dir / excel_filename
        
        self.batch_size = batch_size
        self.pending_data: List[Dict[str, Any]] = []
        
        # Track if CSV file exists (for header handling)
        self.csv_exists = self.csv_path.exists()
        self.excel_exists = self.excel_path.exists()
        
        logger.info(
            f"IncrementalWriter initialized: batch_size={batch_size}, "
            f"CSV exists={self.csv_exists}, Excel exists={self.excel_exists}"
        )
    
    def add_data(self, data: List[Dict[str, Any]]) -> None:
        """Add data to pending batch.
        
        Args:
            data: List of data dictionaries to add
        """
        self.pending_data.extend(data)
        
        # Write batch if threshold reached
        if len(self.pending_data) >= self.batch_size:
            self.flush()
    
    def flush(self) -> None:
        """Force write all pending data immediately."""
        if not self.pending_data:
            return
        
        try:
            # Write to CSV
            self._append_csv(self.pending_data)
            
            # Write to Excel
            self._append_excel(self.pending_data)
            
            logger.debug(f"Flushed {len(self.pending_data)} rows to CSV and Excel")
            self.pending_data = []
            
        except Exception as e:
            logger.error(f"Error flushing data: {e}", exc_info=True)
            # Keep pending data for retry
            raise
    
    def _append_csv(self, data: List[Dict[str, Any]]) -> None:
        """Append data to CSV file.
        
        Args:
            data: List of data dictionaries to append
        """
        if not data:
            return
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Reorder columns to match expected order
        column_order = [field for field, _ in self.COLUMNS]
        
        # Ensure all columns exist (add missing ones with default values)
        for col in column_order:
            if col not in df.columns:
                if col.endswith('_g') or col == 'calories' or col == 'measurement_value':
                    df[col] = 0.0
                elif col in ['food_name', 'measurement_unit']:
                    df[col] = ""
                else:
                    df[col] = None
        
        df = df.reindex(columns=column_order)
        
        # Check if CSV exists and has different columns - if so, we need to update header
        if self.csv_exists:
            try:
                # Read existing CSV to check columns
                existing_df = pd.read_csv(self.csv_path, nrows=0, encoding='utf-8-sig')
                existing_cols = list(existing_df.columns)
                new_cols = list(df.columns)
                
                # If columns don't match, we need to add missing columns to existing data
                if set(existing_cols) != set(new_cols):
                    logger.warning(f"Column mismatch detected. Existing: {existing_cols}, New: {new_cols}")
                    # For now, ensure new data has all required columns
                    # The CSV will have inconsistent columns, but we'll fix it later
                    pass
            except Exception as e:
                logger.debug(f"Could not check existing CSV columns: {e}")
        
        # Append to CSV (with header only if file doesn't exist)
        df.to_csv(
            self.csv_path,
            mode='a' if self.csv_exists else 'w',
            header=not self.csv_exists,
            index=False,
            encoding="utf-8-sig"
        )
        
        # Mark as existing after first write
        if not self.csv_exists:
            self.csv_exists = True
        
        logger.debug(f"Appended {len(data)} rows to CSV: {self.csv_path}")
    
    def _append_excel(self, data: List[Dict[str, Any]]) -> None:
        """Append data to Excel file.
        
        Args:
            data: List of data dictionaries to append
        """
        if not data:
            return
        
        # Load existing workbook or create new
        if self.excel_exists:
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["Nutritional Data"]
                next_row = ws.max_row + 1
            except Exception as e:
                logger.warning(f"Error loading existing Excel file: {e}. Creating new file.")
                wb = None
                ws = None
                next_row = 2
        else:
            wb = None
            ws = None
            next_row = 2
        
        # Create new workbook if needed
        if wb is None:
            wb = Workbook()
            ws = wb.active
            ws.title = "Nutritional Data"
            next_row = 2
            
            # Write headers
            for col_idx, (field, header) in enumerate(self.COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = self.ALIGNMENT_CENTER
                cell.border = self.BORDER
        
        # Write data rows
        for row_data in data:
            for col_idx, (field, _) in enumerate(self.COLUMNS, start=1):
                value = row_data.get(field)
                # Handle missing numeric fields (set to 0.0)
                if field.endswith('_g') or field == 'calories' or field == 'measurement_value':
                    if value is None:
                        value = 0.0
                elif value is None:
                    value = ""
                
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.border = self.BORDER
                
                # Set alignment
                if field in ["food_name", "measurement_unit"]:
                    cell.alignment = self.ALIGNMENT_LEFT
                else:
                    cell.alignment = self.ALIGNMENT_CENTER
                    # Format numeric cells
                    if isinstance(value, (int, float)) and field != "food_id":
                        cell.number_format = '0.0'
            
            next_row += 1
        
        # Auto-adjust column widths (only for new columns or if needed)
        for col_idx, (field, header) in enumerate(self.COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            
            # Calculate max width from all rows
            max_length = len(header)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                if row[0].value:
                    max_length = max(max_length, len(str(row[0].value)))
            
            # Set width with padding
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Atomic save: write to temp file first, then replace
        try:
            temp_file = tempfile.NamedTemporaryFile(
                mode='wb',
                delete=False,
                suffix='.xlsx',
                dir=self.output_dir
            )
            temp_path = Path(temp_file.name)
            temp_file.close()
            
            wb.save(temp_path)
            
            # Atomic replace
            import os
            os.replace(temp_path, self.excel_path)
            
            # Mark as existing after first save
            if not self.excel_exists:
                self.excel_exists = True
            
            logger.debug(f"Appended {len(data)} rows to Excel: {self.excel_path}")
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}", exc_info=True)
            # Clean up temp file if it exists
            if 'temp_path' in locals() and temp_path.exists():
                try:
                    temp_path.unlink()
                except:
                    pass
            raise
    
    def finalize(self) -> None:
        """Write any remaining pending data and finalize files."""
        # Flush any remaining data
        if self.pending_data:
            self.flush()
        
        # Update summary sheet in Excel if file exists
        if self.excel_exists:
            try:
                self._update_summary_sheet()
            except Exception as e:
                logger.warning(f"Could not update summary sheet: {e}")
    
    def _update_summary_sheet(self) -> None:
        """Update or create summary sheet in Excel file."""
        try:
            wb = load_workbook(self.excel_path)
            
            # Remove existing summary sheet if it exists
            if "Summary" in wb.sheetnames:
                wb.remove(wb["Summary"])
            
            # Read all data from main sheet
            ws_data = wb["Nutritional Data"]
            data = []
            for row in ws_data.iter_rows(min_row=2, values_only=False):
                row_dict = {}
                for col_idx, (field, _) in enumerate(self.COLUMNS, start=1):
                    cell = row[col_idx - 1]
                    row_dict[field] = cell.value
                data.append(row_dict)
            
            # Create summary sheet
            ws = wb.create_sheet("Summary", 0)
            
            # Calculate statistics
            unique_foods = len(set(row.get("food_id") for row in data if row.get("food_id")))
            total_rows = len(data)
            completion_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Get measurement unit distribution
            measurement_counts = {}
            for row in data:
                unit = row.get("measurement_unit", "Unknown")
                measurement_counts[unit] = measurement_counts.get(unit, 0) + 1
            
            # Write summary
            summary_data = [
                ["Mankan.me Nutritional Database - Summary", ""],
                ["", ""],
                ["Completion Date", completion_date],
                ["Total Food Items", unique_foods],
                ["Total Data Rows", total_rows],
                ["Average Measurements per Food", f"{total_rows / unique_foods:.2f}" if unique_foods > 0 else "0"],
                ["", ""],
                ["Measurement Unit Distribution", ""],
            ]
            
            # Sort measurement units by count
            sorted_units = sorted(measurement_counts.items(), key=lambda x: x[1], reverse=True)
            for unit, count in sorted_units[:10]:  # Top 10
                summary_data.append([unit, count])
            
            # Write to sheet
            for row_idx, row_data in enumerate(summary_data, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # Title row
                        cell.font = Font(bold=True, size=14)
                    elif row_idx <= 7:  # Stats rows
                        if col_idx == 1:
                            cell.font = Font(bold=True)
            
            # Auto-adjust column widths
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 20
            
            # Save workbook
            wb.save(self.excel_path)
            logger.debug("Summary sheet updated")
            
        except Exception as e:
            logger.error(f"Error updating summary sheet: {e}", exc_info=True)
            raise

