"""Create styled Excel file from fruits_temp.csv for visual comparison."""

import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from src.logger_config import setup_logger

logger = setup_logger()


def create_styled_fruit_excel():
    """Create styled Excel file from fruits_temp.csv."""
    csv_path = Path("output/fruits_temp.csv")
    output_path = Path("output/fruits_temp_styled.xlsx")
    
    if not csv_path.exists():
        logger.error(f"CSV file not found: {csv_path}")
        logger.error("Please run scripts/scrape_fruits_standalone.py first")
        return False
    
    logger.info(f"Reading fruit data from: {csv_path}")
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        logger.info(f"Loaded {len(df)} rows from CSV")
    except Exception as e:
        logger.error(f"Error reading CSV: {e}")
        return False
    
    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fruit Data"
    
    # Define headers (Persian and English)
    headers = [
        ("ردیف", "Row"),  # Row number
        ("نام میوه", "Fruit Name"),  # نام
        ("واحد", "Unit"),  # واحد (should be 100 گرم)
        ("کالری", "Calories"),  # کالری
        ("چربی", "Fat"),  # چربی (should be 0)
        ("پروتئین", "Protein"),  # پروتئین (should be 0)
        ("کربوهیدرات", "Carbohydrates"),  # کربوهیدرات (should be 0)
        ("فیبر", "Fiber"),  # فیبر
        ("قند", "Sugar"),  # قند
        ("نمک", "Salt"),  # نمک (should be 0)
    ]
    
    # Write headers
    for col_idx, (persian, english) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = persian
        cell.fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # Set header row height
    ws.row_dimensions[1].height = 30
    
    # Write data rows
    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        # Column 1: Row number
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        
        # Column 2: Fruit name (نام میوه)
        ws.cell(row=row_idx, column=2, value=row_data.get('food_name', ''))
        
        # Column 3: Unit (واحد)
        ws.cell(row=row_idx, column=3, value=row_data.get('measurement_unit', ''))
        
        # Column 4: Calories (کالری)
        calories = row_data.get('calories', 0)
        ws.cell(row=row_idx, column=4, value=round(calories, 1) if pd.notna(calories) else 0)
        
        # Column 5: Fat (چربی) - should be 0
        fat = row_data.get('fat_g', 0)
        ws.cell(row=row_idx, column=5, value=round(fat, 1) if pd.notna(fat) else 0)
        
        # Column 6: Protein (پروتئین) - should be 0
        protein = row_data.get('protein_g', 0)
        ws.cell(row=row_idx, column=6, value=round(protein, 1) if pd.notna(protein) else 0)
        
        # Column 7: Carbohydrates (کربوهیدرات) - should be 0
        carbs = row_data.get('carbs_g', 0)
        ws.cell(row=row_idx, column=7, value=round(carbs, 1) if pd.notna(carbs) else 0)
        
        # Column 8: Fiber (فیبر)
        fiber = row_data.get('fiber_g', 0)
        ws.cell(row=row_idx, column=8, value=round(fiber, 1) if pd.notna(fiber) else 0)
        
        # Column 9: Sugar (قند)
        sugar = row_data.get('sugar_g', 0)
        ws.cell(row=row_idx, column=9, value=round(sugar, 1) if pd.notna(sugar) else 0)
        
        # Column 10: Salt (نمک) - should be 0
        salt = row_data.get('salt_g', 0)
        ws.cell(row=row_idx, column=10, value=round(salt, 1) if pd.notna(salt) else 0)
        
        # Apply borders and alignment to all cells in the row
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx == 1:  # Row number - center
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:  # Fruit name - right align (Persian text)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 3:  # Unit - right align
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:  # Numeric columns - center
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    column_widths = {
        1: 8,   # Row number
        2: 35,  # Fruit name
        3: 20,  # Unit
        4: 12,  # Calories
        5: 10,  # Fat
        6: 12,  # Protein
        7: 15,  # Carbohydrates
        8: 10,  # Fiber
        9: 10,  # Sugar
        10: 10,  # Salt
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    logger.info(f"Saving styled Excel to: {output_path}")
    try:
        wb.save(output_path)
        logger.info(f"✓ Styled Excel file created: {output_path}")
        logger.info(f"  Total rows: {len(df)}")
        return True
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        return False


def main():
    """Main function."""
    logger.info("=" * 60)
    logger.info("Create Styled Fruit Excel")
    logger.info("=" * 60)
    
    success = create_styled_fruit_excel()
    
    if success:
        logger.info("")
        logger.info("=" * 60)
        logger.info("✓ Complete!")
        logger.info("=" * 60)
        sys.exit(0)
    else:
        logger.error("")
        logger.error("Failed to create styled Excel file")
        sys.exit(1)


if __name__ == "__main__":
    main()

