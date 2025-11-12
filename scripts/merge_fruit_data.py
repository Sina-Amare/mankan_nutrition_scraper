"""Merge verified fruit data to the end of main CSV/Excel files."""

import sys
import shutil
from pathlib import Path
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl import load_workbook
from src.logger_config import setup_logger

logger = setup_logger()


def merge_fruit_data():
    """Merge verified fruit data to the end of main data files."""
    logger.info("=" * 60)
    logger.info("Merge Fruit Data to Main Files")
    logger.info("=" * 60)
    
    # File paths
    temp_csv = Path("output/fruits_temp.csv")
    main_csv = Path("output/mankan_nutritional_data.csv")
    main_excel = Path("output/mankan_nutritional_data.xlsx")
    
    # Check if temporary fruit file exists
    if not temp_csv.exists():
        logger.error(f"Temporary fruit CSV not found: {temp_csv}")
        logger.error("Please run scripts/scrape_fruits_standalone.py first")
        return False
    
    # Check if main CSV exists
    if not main_csv.exists():
        logger.error(f"Main CSV not found: {main_csv}")
        logger.error("Please ensure main data file exists")
        return False
    
    # Create backup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_csv = main_csv.with_suffix(f'.csv.backup_{timestamp}')
    logger.info(f"Creating backup: {backup_csv}")
    shutil.copy2(main_csv, backup_csv)
    
    if main_excel.exists():
        backup_excel = main_excel.with_suffix(f'.xlsx.backup_{timestamp}')
        logger.info(f"Creating backup: {backup_excel}")
        shutil.copy2(main_excel, backup_excel)
    
    # Read fruit data
    logger.info("")
    logger.info("Reading fruit data...")
    try:
        fruit_df = pd.read_csv(temp_csv, encoding='utf-8-sig')
        logger.info(f"  Fruit rows: {len(fruit_df)}")
    except Exception as e:
        logger.error(f"Error reading fruit CSV: {e}")
        return False
    
    # Read main data
    logger.info("Reading main data...")
    try:
        main_df = pd.read_csv(main_csv, encoding='utf-8-sig')
        logger.info(f"  Main rows: {len(main_df)}")
    except Exception as e:
        logger.error(f"Error reading main CSV: {e}")
        return False
    
    # Check for duplicate fruit IDs in main data
    logger.info("")
    logger.info("Checking for duplicate fruit IDs in main data...")
    fruit_ids_in_main = set(main_df['food_id'].unique())
    fruit_ids_to_add = set(fruit_df['food_id'].unique())
    duplicates = fruit_ids_in_main.intersection(fruit_ids_to_add)
    
    if duplicates:
        logger.warning(f"⚠ Found {len(duplicates)} fruit IDs already in main data:")
        logger.warning(f"   Duplicate IDs: {sorted(list(duplicates))[:20]}{'...' if len(duplicates) > 20 else ''}")
        response = input("Do you want to remove existing fruit entries and replace them? (yes/no): ")
        if response.lower() in ['yes', 'y']:
            logger.info("Removing existing fruit entries from main data...")
            main_df = main_df[~main_df['food_id'].isin(duplicates)]
            logger.info(f"  Removed {len(duplicates)} existing fruit entries")
        else:
            logger.info("Skipping duplicate fruit IDs (not adding them)")
            fruit_df = fruit_df[~fruit_df['food_id'].isin(duplicates)]
            logger.info(f"  Will add {len(fruit_df)} new fruit rows")
    
    # Ensure column order matches
    logger.info("")
    logger.info("Ensuring column order matches...")
    main_columns = main_df.columns.tolist()
    fruit_columns = fruit_df.columns.tolist()
    
    # Reorder fruit_df columns to match main_df
    fruit_df = fruit_df[main_columns]
    
    # Append fruit data to the END of main data
    logger.info("")
    logger.info("Appending fruit data to the end of main data...")
    merged_df = pd.concat([main_df, fruit_df], ignore_index=True)
    logger.info(f"  Total rows after merge: {len(merged_df)}")
    logger.info(f"  Added {len(fruit_df)} fruit rows")
    
    # Save merged CSV
    logger.info("")
    logger.info("Saving merged CSV...")
    try:
        merged_df.to_csv(main_csv, index=False, encoding="utf-8-sig")
        logger.info(f"✓ Saved: {main_csv}")
    except Exception as e:
        logger.error(f"Error saving merged CSV: {e}")
        return False
    
    # Update Excel file
    if main_excel.exists():
        logger.info("Updating Excel file...")
        try:
            # Read existing Excel
            wb = load_workbook(main_excel)
            ws = wb.active
            
            # Clear existing data (keep header)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            
            # Write merged data
            from openpyxl.utils.dataframe import dataframe_to_rows
            for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            wb.save(main_excel)
            logger.info(f"✓ Updated: {main_excel}")
        except Exception as e:
            logger.warning(f"⚠ Error updating Excel file: {e}")
            logger.warning("  CSV file was updated successfully. Excel can be regenerated.")
    else:
        logger.warning(f"⚠ Excel file not found: {main_excel}")
        logger.warning("  CSV file was updated successfully.")
    
    # Summary
    logger.info("")
    logger.info("=" * 60)
    logger.info("✓ MERGE COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Main data rows: {len(main_df)}")
    logger.info(f"Fruit rows added: {len(fruit_df)}")
    logger.info(f"Total rows after merge: {len(merged_df)}")
    logger.info("")
    logger.info(f"Backup files created:")
    logger.info(f"  - {backup_csv}")
    if main_excel.exists():
        logger.info(f"  - {backup_excel}")
    logger.info("")
    logger.info("Next step: Run scripts/create_styled_excel.py to regenerate styled Excel")
    logger.info("=" * 60)
    
    return True


def main():
    """Main merge function."""
    try:
        success = merge_fruit_data()
        if success:
            sys.exit(0)
        else:
            sys.exit(1)
    except KeyboardInterrupt:
        logger.warning("Merge interrupted by user.")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Fatal error during merge: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()

