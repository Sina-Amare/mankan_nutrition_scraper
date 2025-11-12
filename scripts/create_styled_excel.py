"""Create a clean, styled Excel file from the CSV data matching the image format."""

import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# Read CSV (from project root)
project_root = Path(__file__).parent.parent
csv_path = project_root / "output/mankan_nutritional_data.csv"
df = pd.read_csv(csv_path, encoding='utf-8-sig')

print(f"Loaded {len(df)} rows from CSV")

# Sort by food_id and measurement_unit for better organization
df = df.sort_values(['food_id', 'measurement_unit'])

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Nutritional Data"

# Persian column headers (matching the image format)
headers = [
    ("ردیف", "Row"),  # Row number
    ("نام غذا", "Food Name"),  # نام
    ("واحد", "Unit"),  # واحد
    ("کالری", "Calories"),  # کالری
    ("چربی", "Fat"),  # چربی
    ("پروتئین", "Protein"),  # پروتئین
    ("کربوهیدرات", "Carbohydrates"),  # کربوهیدرات
    ("فیبر", "Fiber"),  # فیبر
    ("قند", "Sugar"),  # قند
    ("نمک", "Salt"),  # نمک
]

# Styling constants
HEADER_FILL = PatternFill(
    start_color="4472C4",
    end_color="4472C4",
    fill_type="solid"
)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)
ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGNMENT_RIGHT = Alignment(horizontal="right", vertical="center")

# Write headers
header_row = 1
for col_idx, (persian_header, _) in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col_idx, value=persian_header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = ALIGNMENT_CENTER
    cell.border = BORDER

# Write data
row_num = 1
current_food_id = None
food_row_counter = {}  # Track row numbers per food item
previous_food_id = None

for idx, row_data in df.iterrows():
    food_id = row_data['food_id']
    food_name = row_data['food_name']
    
    # Track row numbers per food item (for grouping)
    if food_id not in food_row_counter:
        food_row_counter[food_id] = 1
    else:
        food_row_counter[food_id] += 1
    
    # Write row
    data_row = header_row + row_num
    
    # Add visual separation between different food items
    if previous_food_id is not None and food_id != previous_food_id:
        # Add a slightly thicker bottom border to separate food groups
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row - 1, column=col_idx)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=Side(style="medium", color="CCCCCC")
            )
    
    previous_food_id = food_id
    
    # Column 1: Row number (ردیف)
    ws.cell(row=data_row, column=1, value=food_row_counter[food_id])
    
    # Column 2: Food Name (نام غذا)
    ws.cell(row=data_row, column=2, value=food_name)
    
    # Column 3: Unit (واحد)
    ws.cell(row=data_row, column=3, value=row_data['measurement_unit'])
    
    # Column 4: Calories (کالری)
    calories = row_data['calories']
    ws.cell(row=data_row, column=4, value=round(calories, 1) if pd.notna(calories) else "")
    
    # Column 5: Fat (چربی)
    fat = row_data['fat_g']
    ws.cell(row=data_row, column=5, value=round(fat, 1) if pd.notna(fat) else "")
    
    # Column 6: Protein (پروتئین)
    protein = row_data['protein_g']
    ws.cell(row=data_row, column=6, value=round(protein, 1) if pd.notna(protein) else "")
    
    # Column 7: Carbohydrates (کربوهیدرات)
    carbs = row_data.get('carbs_g', 0)
    ws.cell(row=data_row, column=7, value=round(carbs, 1) if pd.notna(carbs) else 0)
    
    # Column 8: Fiber (فیبر)
    fiber = row_data.get('fiber_g', 0)
    ws.cell(row=data_row, column=8, value=round(fiber, 1) if pd.notna(fiber) else 0)
    
    # Column 9: Sugar (قند)
    sugar = row_data.get('sugar_g', 0)
    ws.cell(row=data_row, column=9, value=round(sugar, 1) if pd.notna(sugar) else 0)
    
    # Column 10: Salt (نمک)
    salt = row_data.get('salt_g', 0)
    ws.cell(row=data_row, column=10, value=round(salt, 1) if pd.notna(salt) else 0)
    
    # Apply styling to all cells in this row (including new columns)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=data_row, column=col_idx)
        cell.border = BORDER
        
        # Alignment based on column
        if col_idx == 1:  # Row number
            cell.alignment = ALIGNMENT_CENTER
        elif col_idx == 2:  # Food name
            cell.alignment = ALIGNMENT_LEFT
        elif col_idx == 3:  # Unit
            cell.alignment = ALIGNMENT_LEFT
        else:  # Numbers
            cell.alignment = ALIGNMENT_CENTER
            cell.number_format = '0.0'
    
    row_num += 1

# Auto-adjust column widths
column_widths = {
    1: 8,   # Row number
    2: 35,  # Food name
    3: 20,  # Unit
    4: 12,  # Calories
    5: 10,  # Fat
    6: 12,  # Protein
    7: 15,  # Carbohydrates
    8: 10,  # Fiber
    9: 10,  # Sugar
    10: 10,  # Salt
}

for col_idx, width in column_widths.items():
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Add alternating row colors for better readability (grouped by food item)
LIGHT_FILL = PatternFill(
    start_color="F8F9FA",
    end_color="F8F9FA",
    fill_type="solid"
)

current_food_id_for_styling = None
row_style_counter = 0

for row_idx in range(2, ws.max_row + 1):
    # Get food_id from the data row (we need to track this)
    food_name_cell = ws.cell(row=row_idx, column=2)
    # We'll apply styling based on position - every other food group gets light fill
    # For simplicity, we'll use row position
    if row_idx % 2 == 0:  # Even rows get light gray
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.fill.start_color.index == "00000000":  # Only if no fill
                cell.fill = LIGHT_FILL

# Save workbook (to project root output directory)
output_path = project_root / "output/mankan_nutritional_data_styled.xlsx"
wb.save(output_path)

print(f"Created styled Excel file: {output_path}")
print(f"Total rows: {len(df)}")
print(f"Total food items: {df['food_id'].nunique()}")

